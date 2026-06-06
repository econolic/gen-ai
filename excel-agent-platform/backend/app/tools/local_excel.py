from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from app.schemas.run import CellUpdate
from app.schemas.workbook import ColumnProfile, SheetProfile, WorkbookProfile


def _json_safe(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def _read_sheet(file_path: str | Path, sheet_name: str | None = None) -> pd.DataFrame:
    df = pd.read_excel(file_path, sheet_name=sheet_name or 0, engine="openpyxl")
    df = df.dropna(how="all")
    df.columns = [str(column).strip() for column in df.columns]
    return df


def profile_workbook(file_path: str | Path) -> WorkbookProfile:
    """Read workbook shape, columns, nulls, and sample rows."""

    path = Path(file_path)
    sheets: list[SheetProfile] = []
    workbook = pd.ExcelFile(path, engine="openpyxl")

    for sheet in workbook.sheet_names:
        df = _read_sheet(path, sheet)
        columns: list[ColumnProfile] = []
        for column in df.columns:
            series = df[column]
            sample_values = [_json_safe(v) for v in series.dropna().head(5).tolist()]
            columns.append(
                ColumnProfile(
                    name=str(column),
                    non_null_count=int(series.notna().sum()),
                    null_count=int(series.isna().sum()),
                    dtype=str(series.dtype),
                    sample_values=sample_values,
                )
            )

        sample_rows = [
            {str(k): _json_safe(v) for k, v in row.items()}
            for row in df.head(5).to_dict(orient="records")
        ]
        sheets.append(
            SheetProfile(
                name=sheet,
                row_count=int(len(df)),
                column_count=int(len(df.columns)),
                columns=columns,
                sample_rows=sample_rows,
            )
        )

    return WorkbookProfile(file_path=str(path), sheets=sheets)


def read_rows(file_path: str | Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    df = _read_sheet(file_path, sheet_name)
    rows = []
    for index, row in df.iterrows():
        rows.append(
            {
                "_row_index": int(index),
                **{str(k): _json_safe(v) for k, v in row.items()},
            }
        )
    return rows


def preview_rows(file_path: str | Path, limit: int = 20, sheet_name: str | None = None) -> list[dict[str, Any]]:
    return read_rows(file_path, sheet_name)[:limit]


def write_enriched_workbook(
    input_path: str | Path,
    output_path: str | Path,
    updates: list[CellUpdate],
    sheet_name: str | None = None,
) -> str:
    """Preserve workbook structure and write only requested target cells."""

    source = Path(input_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    header_cells = list(worksheet[1])
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]

    for update in updates:
        if update.target_column not in headers:
            headers.append(update.target_column)
            worksheet.cell(row=1, column=len(headers), value=update.target_column)
        column_number = headers.index(update.target_column) + 1
        worksheet.cell(row=update.row_index + 2, column=column_number, value=update.value)

    workbook.save(output)
    return str(output)
