import json
from pathlib import Path

from openpyxl import load_workbook

from app.config import get_settings
from app.process import process_excel


ROOT_DIR = Path(__file__).resolve().parents[2]
FIXTURES_DIR = ROOT_DIR / "tests" / "fixtures"
GOLDEN_DIR = ROOT_DIR / "tests" / "golden"


def _reset_settings(monkeypatch, tmp_path):
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.setenv("OFFLINE_DEMO_SEED_FIRST", "true")
    get_settings.cache_clear()


def _within_tolerance(value, expected, tolerance_pct):
    tolerance = abs(expected) * tolerance_pct / 100
    return abs(float(value) - expected) <= tolerance


def test_capitals_golden_accuracy(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    expected = json.loads((GOLDEN_DIR / "capitals_expected.json").read_text(encoding="utf-8"))
    result = process_excel(
        str(FIXTURES_DIR / "capitals.xlsx"),
        "find the straight-line distance between the capitals in kilometers for the column distance",
    )

    sheet = load_workbook(result.output_path, data_only=True).active
    passed = 0
    covered = 0
    for row in range(2, sheet.max_row + 1):
        key = "|".join(str(sheet.cell(row=row, column=col).value) for col in range(1, 5))
        if key not in expected:
            continue
        covered += 1
        spec = expected[key]
        value = sheet.cell(row=row, column=5).value
        passed += int(value is not None and _within_tolerance(value, spec["expected"], spec["tolerance_pct"]))

    assert covered == len(expected)
    assert passed / covered >= 0.95


def test_mountains_golden_accuracy(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    expected = json.loads((GOLDEN_DIR / "mountains_expected.json").read_text(encoding="utf-8"))
    result = process_excel(
        str(FIXTURES_DIR / "mountains.xlsx"),
        "add the height of the mountains in meters to the column height",
    )

    sheet = load_workbook(result.output_path, data_only=True).active
    passed = 0
    covered = 0
    for row in range(2, sheet.max_row + 1):
        key = f"{sheet.cell(row=row, column=1).value}|{sheet.cell(row=row, column=2).value}"
        if key not in expected:
            continue
        covered += 1
        spec = expected[key]
        value = sheet.cell(row=row, column=3).value
        passed += int(value is not None and _within_tolerance(value, spec["expected"], spec["tolerance_pct"]))

    assert covered == len(expected)
    assert passed / covered >= 0.95


def test_final_project_mountains_dataset(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    expected = {
        "Everest|Nepal": 8848,
        "Mont Blanc|France": 4805,
        "Denali|USA": 6190,
        "Kilimanjaro|Tanzania": 5895,
    }
    result = process_excel(
        str(FIXTURES_DIR / "mountains_final_project.xlsx"),
        "add the height of the mountains in meters to the column height",
    )

    sheet = load_workbook(result.output_path, data_only=True).active
    covered = 0
    for row in range(2, sheet.max_row + 1):
        key = f"{sheet.cell(row=row, column=1).value}|{sheet.cell(row=row, column=2).value}"
        expected_value = expected[key]
        value = sheet.cell(row=row, column=3).value
        covered += int(value is not None and _within_tolerance(value, expected_value, 2))

    assert covered == len(expected)
