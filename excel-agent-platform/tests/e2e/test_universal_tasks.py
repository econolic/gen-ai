from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.config import get_settings
from app.process import process_excel
from app.schemas.evidence import Evidence, FactRequest, FactResult
from app.services import mcp_gateway


def _reset_settings(monkeypatch, tmp_path):
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.setenv("OFFLINE_DEMO_SEED_FIRST", "true")
    get_settings.cache_clear()


def _workbook(path: Path, rows: list[list]):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_formula_with_plain_columns(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    path = tmp_path / "formula.xlsx"
    _workbook(path, [["A", "B", "C"], [10, 5, None], [20, 7, None]])

    result = process_excel(str(path), "обчисли A+B і збережи результат у колонку C")
    sheet = load_workbook(result.output_path, data_only=True).active

    assert [sheet["C2"].value, sheet["C3"].value] == [15, 27]
    assert result.plan.operations[0].type == "formula"


def test_formula_with_spaced_columns(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    path = tmp_path / "margin.xlsx"
    _workbook(path, [["Sales Amount", "Cost Amount", "Margin"], [1000, 700, None]])

    result = process_excel(str(path), "calculate Margin as Sales Amount minus Cost Amount")
    sheet = load_workbook(result.output_path, data_only=True).active

    assert sheet["C2"].value == 300
    assert "col('Sales Amount')" in result.plan.operations[0].expression


def test_group_share(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    path = tmp_path / "share.xlsx"
    _workbook(path, [["Product", "Sales", "Share"], ["A", 100, None], ["B", 300, None]])

    result = process_excel(str(path), "обчисли частку кожного товару в загальних продажах")
    sheet = load_workbook(result.output_path, data_only=True).active

    assert [sheet["C2"].value, sheet["C3"].value] == [0.25, 0.75]
    assert result.plan.operations[0].type == "group_share"


def test_generic_lookup(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    path = tmp_path / "population.xlsx"
    _workbook(path, [["City", "Country", "population"], ["Paris", "France", None]])

    async def fake_lookup(request: FactRequest) -> FactResult:
        return FactResult(
            request=request,
            value=2148000,
            unit="people",
            confidence=0.9,
            evidence=[Evidence(kind="wikidata", title="Paris", confidence=0.9)],
        )

    monkeypatch.setattr(mcp_gateway, "lookup_fact", fake_lookup)
    result = process_excel(str(path), "додай населення міст у колонку population")
    sheet = load_workbook(result.output_path, data_only=True).active

    assert sheet["C2"].value == 2148000
    assert result.plan.operations[0].attribute == "population"


def test_hybrid_lookup_then_formula(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    path = tmp_path / "hybrid.xlsx"
    _workbook(path, [["City", "Country", "Sales", "Population", "Sales_per_capita"], ["Paris", "France", 1000000, None, None]])

    async def fake_lookup(request: FactRequest) -> FactResult:
        return FactResult(
            request=request,
            value=2000000,
            unit="people",
            confidence=0.9,
            evidence=[Evidence(kind="wikidata", title="Paris", confidence=0.9)],
        )

    monkeypatch.setattr(mcp_gateway, "lookup_fact", fake_lookup)
    result = process_excel(str(path), "відшукай населення міста і розрахуй продажі на душу населення")
    sheet = load_workbook(result.output_path, data_only=True).active

    assert sheet["D2"].value == 2000000
    assert sheet["E2"].value == 0.5
    assert len(result.plan.operations) == 2
