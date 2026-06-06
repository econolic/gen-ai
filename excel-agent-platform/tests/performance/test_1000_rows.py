import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.config import get_settings
from app.process import process_excel


def _reset_settings(monkeypatch, tmp_path):
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    get_settings.cache_clear()


def _workbook(path: Path, rows: list[list]):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_formula_1000_rows_under_5_seconds(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    path = tmp_path / "formula_1000.xlsx"
    _workbook(path, [["A", "B", "C"], *[[index, index + 1, None] for index in range(1000)]])

    started = time.perf_counter()
    result = process_excel(str(path), "обчисли A+B і збережи результат у колонку C")
    elapsed = time.perf_counter() - started
    sheet = load_workbook(result.output_path, data_only=True).active

    assert elapsed < 5
    assert sheet["C1001"].value == 1999
    assert result.plan.estimated_external_calls == 0


def test_group_share_1000_rows_under_5_seconds(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    path = tmp_path / "share_1000.xlsx"
    _workbook(path, [["Product", "Sales", "Share"], *[[f"P{index}", 1, None] for index in range(1000)]])

    started = time.perf_counter()
    result = process_excel(str(path), "обчисли частку кожного товару в загальних продажах")
    elapsed = time.perf_counter() - started
    sheet = load_workbook(result.output_path, data_only=True).active

    assert elapsed < 5
    assert sheet["C2"].value == 0.001
    assert result.plan.estimated_external_calls == 0
