import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl import Workbook

from app.config import get_settings
from app.process import process_excel
from app.services import operation_resolver
from app.services.operation_resolver import OperationResolution


ROOT_DIR = Path(__file__).resolve().parents[2]
FIXTURES_DIR = ROOT_DIR / "tests" / "fixtures"


def _reset_settings(monkeypatch, tmp_path):
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    get_settings.cache_clear()


def test_process_capitals_enriches_distance(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    result = process_excel(
        str(FIXTURES_DIR / "capitals.xlsx"),
        "find the straight-line distance between the capitals in kilometers for the column distance",
    )
    workbook = load_workbook(result.output_path, data_only=True)
    sheet = workbook.active
    values = [sheet[f"E{row}"].value for row in range(2, 12)]
    assert all(isinstance(value, int | float) for value in values)
    assert 1900 <= values[0] <= 2100


def test_process_mountains_enriches_height(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    result = process_excel(
        str(FIXTURES_DIR / "mountains.xlsx"),
        "add the height of the mountains in meters to the column height",
    )
    workbook = load_workbook(result.output_path, data_only=True)
    sheet = workbook.active
    values = [sheet[f"C{row}"].value for row in range(2, 12)]
    assert values[0] == 8848
    assert values[1] == 8611
    assert all(isinstance(value, int | float) for value in values)

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert report["plan"]["operation"] == "lookup"
    assert len(report["updates"]) == 10


def test_process_row_operation_task(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)
    input_path = tmp_path / "math.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["A", "B", "Operation", "Value"])
    sheet.append([10, 20, "addition", None])
    sheet.append([20, 30, "subtraction", None])
    sheet.append([25.5, 54, "multiplication", None])
    sheet.append([100, 4, "division", None])
    workbook.save(input_path)

    result = process_excel(
        str(input_path),
        "calculate the values of columns A and B according to the operation from column Operation "
        "and save the result in column Value",
    )

    output = load_workbook(result.output_path, data_only=True)
    sheet = output.active
    assert sheet["D2"].value == 30
    assert sheet["D3"].value == -10
    assert sheet["D4"].value == 1377
    assert sheet["D5"].value == 25
    assert result.plan.operation == "formula"


def test_process_row_operation_uses_llm_fallback_for_unknown_labels(monkeypatch, tmp_path):
    _reset_settings(monkeypatch, tmp_path)

    async def fake_classify(labels):
        assert labels == ["suma"]
        return {
            "suma": OperationResolution(
                symbol="+",
                confidence=0.93,
                source="llm",
                reason="Spanish addition label",
            )
        }

    monkeypatch.setattr(operation_resolver, "_classify_with_openrouter", fake_classify)

    input_path = tmp_path / "math_multilang.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["A", "B", "Operation", "Value"])
    sheet.append([10, 20, "suma", None])
    workbook.save(input_path)

    result = process_excel(
        str(input_path),
        "calculate the values of columns A and B according to the operation from column Operation "
        "and save the result in column Value",
    )

    output = load_workbook(result.output_path, data_only=True)
    sheet = output.active
    assert sheet["D2"].value == 30
    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert report["updates"][0]["evidence"][0]["metadata"]["resolver_source"] == "llm"


def test_process_large_source_backed_workbook_uses_graph_fanout(monkeypatch, tmp_path):
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.setenv("GRAPH_FANOUT_THRESHOLD", "10")
    monkeypatch.setenv("GRAPH_CHUNK_SIZE", "7")
    monkeypatch.setenv("GRAPH_FANOUT_CONCURRENCY", "2")
    get_settings.cache_clear()

    input_path = tmp_path / "large_mountains.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Mountain", "Country", "height"])
    for _ in range(25):
        sheet.append(["Mount Everest", "Nepal/China", None])
    workbook.save(input_path)

    result = process_excel(
        str(input_path),
        "add the height of the mountains in meters to the column height",
    )

    output = load_workbook(result.output_path, data_only=True)
    sheet = output.active
    values = [sheet[f"C{row}"].value for row in range(2, 27)]
    assert values == [8848] * 25

    report = json.loads(Path(result.report_path).read_text(encoding="utf-8"))
    assert report["fanout"]["enabled"] is True
    assert report["fanout"]["chunk_count"] == 4
    assert sorted(report["processed_chunks"]) == [0, 1, 2, 3]
